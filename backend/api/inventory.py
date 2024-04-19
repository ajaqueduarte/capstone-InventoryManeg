from fastapi import FastAPI, HTTPException
from openpyxl import load_workbook
from typing import List
from pathlib import Path

app = FastAPI()

# Load the Excel file
file_path = Path('./inventory/Inventory_items.xlsx')

try:
    workbook = load_workbook(file_path)
except FileNotFoundError:
    raise FileNotFoundError(f"File '{file_path}' not found.")
except Exception as e:
    raise Exception(f"An error occurred while loading the workbook: {e}")

# Assuming the sheet name is 'Sheet1', you can adjust it accordingly
inventory_sheet = workbook['Sheet1']

# Function to convert Excel sheet to list of dictionaries
def sheet_to_json(sheet):
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

# Read data from Excel sheet and convert to JSON
inventory_data = sheet_to_json(inventory_sheet)

# Function to save data to Excel file
def save_to_excel(data: List[dict], sheet):
    for row_index, row_data in enumerate(data, start=2):
        for column_index, (key, value) in enumerate(row_data.items(), start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

# Endpoint to get all inventory items
@app.get("/inventory/", response_model=List[dict])
async def get_inventory():
    print("Retrieving inventory data...")
    return inventory_data

# Endpoint to update an inventory item
@app.put("/inventory/{item_id}", response_model=dict)
async def update_inventory(item_id: int, item: dict):
    print(f"Updating inventory item with ID: {item_id}")
    for inventory_item in inventory_data:
        if inventory_item['ID'] == item_id:
            inventory_item.update(item)
            save_to_excel(inventory_data, inventory_sheet)
            workbook.save('./inventory/Inventory_items.xlsx')
            return {"message": "Inventory item updated successfully"}
    raise HTTPException(status_code=404, detail="Inventory item not found")

# Endpoint to add a new inventory item
@app.post("/inventory/", response_model=dict)
async def add_inventory(item: dict):
    print("Adding a new inventory item...")
    new_id = max(item['ID'] for item in inventory_data) + 1
    item['ID'] = new_id
    inventory_data.append(item)
    save_to_excel(inventory_data, inventory_sheet)
    workbook.save('./inventory/Inventory_items.xlsx')
    return {"message": "Inventory item added successfully"}

# Endpoint to delete an inventory item by ID
@app.delete("/inventory/{item_id}", response_model=dict)
async def delete_inventory(item_id: int):
    print(f"Deleting inventory item with ID: {item_id}")
    for index, inventory_item in enumerate(inventory_data):
        if inventory_item['ID'] == item_id:
            del inventory_data[index]
            save_to_excel(inventory_data, inventory_sheet)
            workbook.save('./inventory/Inventory_items.xlsx')
            return {"message": "Inventory item deleted successfully"}
    raise HTTPException(status_code=404, detail="Inventory item not found")
