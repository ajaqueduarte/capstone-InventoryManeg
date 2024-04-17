import os
import pandas as pd
import openpyxl
from datetime import datetime
import os


# Directory paths
script_directory = os.path.dirname(os.path.abspath(__file__))
sales_mix_directory = os.path.join(script_directory, 'sales_mix')
# Dictionary to hold the data
date_product_data = {}

# Function to process each Excel file
def process_excel_file(file_path, date_product_data):
    # Load the workbook and get the active sheet
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Find the header row. Assuming 'Name' and 'Quantity Sold' could be in any column.
    header_row_idx = None
    for rowIndex, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if 'Name' in row and 'Quantity Sold' in row:
            header_row_idx = rowIndex
            headers = [cell for cell in row if cell is not None]
            name_idx = headers.index('Name')
            quantity_sold_idx = headers.index('Quantity Sold')
            break

    if header_row_idx is None:
        raise ValueError(f"'Name' and 'Quantity Sold' columns not found in the file: {file_name}")

    # Parse the date from the filename
    date_part = ' '.join(os.path.basename(file_path).split(' ')[2:4]).replace(' - Copy', '')

    # Initialize the dictionary for the day if not already done
    if date_part not in date_product_data:
        date_product_data[date_part] = {}

    # Extracting data for each product
    for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
        product_name = row[name_idx]
        quantity_sold = row[quantity_sold_idx] or 0  # Use 0 if None

        # Skip products based on names to remove or containing specific patterns
        names_to_remove = [
            "Total", "Smoothie", "Classic", "Hi Protein", "Spirit", "Superfood Plus",
            "Superfood", "Refresh", "Combo", "Regular Combo", "Snack Combo",
            "Ingredients (Smoothie)", "NO BOOSTER", "No Froyo", "No Yogurt","Pirates Nectar - R","Pirates Nectar - S","Hawaiian Sunset - R","Hawaiian Sunset - S","Acai Avalanche - R",
            "Secret/Feature", "Specialty", "Shot", "Condiments (Booster Ball)","Lemon Ginger Tumeric & Coconut S","Lemon Ginger Tumeric and Coconut","Ocean Mist - S","Ocean Mist - R","Gingerbread - R","Wildberry Rush - R","Mangosicle - R","Mangosicle - S ","Oatrageous Mocha - S","Oatrageous Mocha - R",
            "Booster Ball", "Booster Blends", "Grilled Fresh", "Lunch", "Breakfast","Macadamia Nut - Booster Ball","Candy Cane - Booster Ball","Birthday Cake - Booster Ball","Apple, Lemon & Ginger Shot - S",
            "Merchandise", "Retail", "Fresh Juice", "Condiments (Fresh Juice)","Egg White & Chorizo Wrap","Egg White & Cheese Wrap","Chocolate Peanut - Protein & Co","Caramel Cashew - Protein & Co","Mango Teazer - TB","Monster - Booster Ball",
            "Instructions", "Condiments (Inst.)", "Grilled Cheese", "Canadian Maple - Booster Ball","Go Mango - R", "Wildberry Rush - S","Chipotle Steak Panini", "Raspberry Rapture - R","Gingerbread - S","Artisan Grilled Cheese",
            "No Raspberry", "Special Prep", "whole wheat", "Split In 2 Cups", "Don't Make", "Almighty Acai Blend", "Condiments (Retail/Merchandise)", "Chipotle"]

        if any(pattern.lower() in product_name.lower() for pattern in ['allergy', 'combo', '$', 'add', 'No']):
            continue
        if product_name in names_to_remove:
            continue

        # Add the quantity sold to the dictionary
        date_product_data[date_part][product_name] = date_product_data[date_part].get(product_name, 0) + quantity_sold

# Process each file
for file_name in os.listdir(sales_mix_directory):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(sales_mix_directory, file_name)
        process_excel_file(file_path, date_product_data)
# Convert the dictionary to a DataFrame 
product_sales_df = pd.DataFrame.from_dict(date_product_data, orient='columns')

# transpose the DataFrame
product_sales_df = product_sales_df.transpose()

# product_sales_df.index = pd.to_datetime(product_sales_df.index + ', 2024', format='%b %d, %Y')


def adjust_year(date):
    # Adjust to 2024 if the month is December
    if date.month == 12:
        return date.replace(year=2023)
    else:
        return date.replace(year=2024)

# Adjust the index for each row in the DataFrame
product_sales_df.index = pd.to_datetime(product_sales_df.index + ', 2024', format='%b %d, %Y', errors='coerce')
product_sales_df.index = product_sales_df.index.map(adjust_year)
product_sales_df.sort_index(inplace=True)


# Convert NaN values to 0
# Replace negative values with 0 and convert types from float to integer
product_sales_df = product_sales_df.fillna(0).clip(lower=0).astype(int)



    # Directory to save the cleaned data file
cleaned_data_directory = os.path.join(script_directory, 'sales_mix_clean_open')

    # Create the directory if it doesn't exist
os.makedirs(cleaned_data_directory, exist_ok=True)

    # Construct the output CSV file path
output_csv_path = os.path.join(cleaned_data_directory, 'sales_mix_cleaned.csv')

    # Save the cleaned data to CSV
product_sales_df.to_csv(output_csv_path, index_label='Date')
    
cleaned_data = product_sales_df

