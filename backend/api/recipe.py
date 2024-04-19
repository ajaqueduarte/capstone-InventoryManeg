from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from typing import List
from pathlib import Path
import json
app = FastAPI()

# Load the Excel file
file_path = Path('./recipes/Recipe.xlsx')

try:
    workbook = load_workbook(file_path)
except FileNotFoundError:
    print(f"File '{file_path}' not found.")
except Exception as e:
    print(f"An error occurred while loading the workbook: {e}")


smoothie_sheet = workbook['smoothie']
juice_sheet = workbook['juice']
food_sheet = workbook['food']


# Function to convert Excel sheet to list of dictionaries
def sheet_to_json(sheet):
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return json.dumps(data)

# Read data from Excel sheets and convert to JSON
smoothie_data = sheet_to_json(smoothie_sheet)
juice_data = sheet_to_json(juice_sheet)
food_data = sheet_to_json(food_sheet)

# Function to save data to Excel file
def save_to_excel(data: List[dict], sheet):
    for row_index, row_data in enumerate(data, start=2):
        for column_index, (key, value) in enumerate(row_data.items(), start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

# Endpoint to get all smoothies
@app.get("/smoothies/")
async def get_smoothies():
    return smoothie_data

# Endpoint to update a smoothie recipe
@app.put("/smoothies/{smoothie_id}")
async def update_smoothie(smoothie_id: int, smoothie: dict):
    for item in smoothie_data:
        if item['id'] == smoothie_id:
            item.update(smoothie)
            save_to_excel(smoothie_data, smoothie_sheet)
            workbook.save('./recipes/Recipe.xlsx')
            return {"message": "Smoothie updated successfully"}
    raise HTTPException(status_code=404, detail="Smoothie not found")

# Endpoint to add a new smoothie recipe
@app.post("/smoothies/")
async def add_smoothie(smoothie: dict):
    new_id = max(item['id'] for item in smoothie_data) + 1
    smoothie['id'] = new_id
    smoothie_data.append(smoothie)
    save_to_excel(smoothie_data, smoothie_sheet)
    workbook.save('./recipes/Recipe.xlsx')
    return {"message": "Smoothie added successfully"}

# Endpoint to delete a smoothie by ID
@app.delete("/smoothies/{smoothie_id}")
async def delete_smoothie(smoothie_id: int):
    for index, item in enumerate(smoothie_data):
        if item['id'] == smoothie_id:
            del smoothie_data[index]
            save_to_excel(smoothie_data, smoothie_sheet)
            workbook.save('./recipes/Recipe.xlsx')
            return {"message": "Smoothie deleted successfully"}
    raise HTTPException(status_code=404, detail="Smoothie not found")


## Juice Endpoints

# Endpoint to get all juices
@app.get("/juices/")
async def get_juices():
    return juice_data

# Endpoint to update a juice recipe
@app.put("/juices/{juice_id}")
async def update_juice(juice_id: int, juice: dict):
    for item in juice_data:
        if item['id'] == juice_id:
            item.update(juice)
            save_to_excel(juice_data, juice_sheet)
            workbook.save('./recipes/Recipe.xlsx')
            return {"message": "Juice updated successfully"}
    raise HTTPException(status_code=404, detail="Juice not found")

# Endpoint to add a new juice recipe
@app.post("/juices/")
async def add_juice(juice: dict):
    new_id = max(item['id'] for item in juice_data) + 1
    juice['id'] = new_id
    juice_data.append(juice)
    save_to_excel(juice_data, juice_sheet)
    workbook.save('./recipes/Recipe.xlsx')
    return {"message": "Juice added successfully"}

# Endpoint to delete a juice by ID
@app.delete("/juices/{juice_id}")
async def delete_juice(juice_id: int):
    for index, item in enumerate(juice_data):
        if item['id'] == juice_id:
            del juice_data[index]
            save_to_excel(juice_data, juice_sheet)
            workbook.save('./recipes/Recipe.xlsx')
            return {"message": "Juice deleted successfully"}
    raise HTTPException(status_code=404, detail="Juice not found")


## Food Endpoints

# Endpoint to get all foods
@app.get("/foods/")
async def get_foods():
    return food_data

# Endpoint to update a food recipe
@app.put("/foods/{food_id}")
async def update_food(food_id: int, food: dict):
    for item in food_data:
        if item['id'] == food_id:
            item.update(food)
            save_to_excel(food_data, food_sheet)
            workbook.save('./recipes/Recipe.xlsx')
            return {"message": "Food updated successfully"}
    raise HTTPException(status_code=404, detail="Food not found")

# Endpoint to add a new food recipe
@app.post("/foods/")
async def add_food(food: dict):
    new_id = max(item['id'] for item in food_data) + 1
    food['id'] = new_id
    food_data.append(food)
    save_to_excel(food_data, food_sheet)
    workbook.save('./recipes/Recipe.xlsx')
    return {"message": "Food added successfully"}

# Endpoint to delete a food by ID
@app.delete("/foods/{food_id}")
async def delete_food(food_id: int):
    for index, item in enumerate(food_data):
        if item['id'] == food_id:
            del food_data[index]
            save_to_excel(food_data, food_sheet)
            workbook.save('./recipes/Recipe.xlsx')
            return {"message": "Food deleted successfully"}
    raise HTTPException(status_code=404, detail="Food not found")
